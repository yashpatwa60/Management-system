from tkinter import *
from tkinter import ttk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import filedialog, Text
import shutil
import os
from functools import partial
from datetime import date  
import pickle
import matplotlib.pyplot as plt
from tkinter import messagebox
# module imported from same directory
from bot import appbot



class application:

    def __init__(self, master):
        self.master = master
        # treeview
        self.tree = ttk.Treeview()
        # menuview
        self.menu = Menu(master)
        master.config(menu = self.menu)
        # notebook
        self.notebook = ttk.Notebook()
        # workbooks
        self.wb = Workbook()
        #time
        self.today = date.today()

        #variables
        self.recent = ''
        self.item_text = ''
        self.filepath = ''
        self.presentbuttons = []
        self.absentbuttons = []
        self.labels = []
        self.row = 0
        self.column = 0
        self.length = 0
        self.n = 0

        self.month_name = 0
        self.sheet_no = 0

        self.tab_frame = Frame(self.notebook, bg= 'white')
        self.tab_frame_dict = {}
        



#   ---------------------- chat-bot --------------------

    def my_bot(self):
        root = Tk()
        root.title("CHATBOT")
        root.geometry('450x600+550+110')
        root.resizable(False, False)
        root.iconbitmap('ai.ico')

        bot = appbot(root)
        bot.doitall()

        root.mainloop()


#   ---------------------- messageboxes ------------------

    def mb_loadprev(self):
        response = messagebox.showerror('Error!', 'first import excel sheet')
        

#   ----------------------- datetime ------------------------
    def get_date(self):
        today = self.today
        date = int(today.strftime("%d"))
        self.recent = today.strftime('%d-%m-%y')
        return date

    def get_month(self):
        today = self.today
        month = int(today.strftime("%m"))
        return month

    def date_format(self, str):                                  
        today = date.today()
        return today.strftime(str)


    def load_datelist(self):
        datelist = []
        try:
            file_name = "sample.pkl"
            open_file = open(file_name, "rb")
            datelist = pickle.load(open_file)
        except:
            print("First timer")

        return datelist

    def append_datelist(self):                                                              # check for same date and change of month
        datelist = self.load_datelist()
        # print(datelist)
        date = self.get_date()
        month = self.get_month()

        if len(datelist) == 0:
                datelist.append((date, month))
                print('flash : first element')
        else:
            for x,y in datelist:
                if x==date and y == month:
                    flag = 0
                    
                else:
                    flag = 1
                    
            if flag == 1:
                for x,y in datelist:
                    if y != month:
                        datelist = []
                        datelist.append((date,month))
                        print('flash : month has been changed')
                        break
                    else:
                        datelist.append((date, month))
                        break
            else:
                print("flash : niether append")
        

        # print(datelist)
        self.length = len(datelist)

        return datelist

    def dumpdatelist(self):
        datelist = self.append_datelist()
        file_name = "sample.pkl"

        open_file = open(file_name, "wb")
        pickle.dump(datelist, open_file)
        open_file.close()       

# ------------------------ Treeview ------------------------------------
 
    def OnDoubleClick(self,event):
        item = self.tree.selection()[0]
        self.item_text = self.tree.item(item, "text")

        if self.item_text in ("Directory", "Files", "final.xlsx","hip.xlsx"):
            pass
        
        else:
            # print(f'Your selected Folder: {self.item_text}')
            self.add_tabs()
            

    def SUB_paths(self, path, parent):
        for p in os.listdir(path):

            abspath = os.path.join(path, p)

            parent_element = self.tree.insert(parent, 'end', text =p, open= False)
            self.tree.bind("<Double-1>", self.OnDoubleClick)

            if os.path.isdir(abspath):
                self.SUB_paths(abspath, parent_element) 

    def Treeview(self):

        path = os.path.join(os.getcwd(),'Files')
        show_path = 'Files'

        self.tree.place(x = 0 , y = 0, height=600 , width = 200)
        self.tree.heading("#0", text = "Directory")

        parent = self.tree.insert('','end', text = show_path, open = True)

        self.SUB_paths(path, parent)   


# _________________________getting directories___________________________


    def get_filebasename(self):                                                                #hip.xlsx
        filebasename = os.path.basename(self.filepath)
        return filebasename

    def input_save_file(self):           
        
        savedfilepath = os.path.join(f"Files\{self.item_text}", self.get_filebasename())
        return savedfilepath

    def output_save_file(self):                                                                 #resolve : first check input file i.e hip.xlsx

        finalfile = os.path.join(f"Files\{self.item_text}", "final.xlsx")
        # finalfile = f"Files\\{self.item_text}\\final.xlsx"
        return finalfile

    def destination_path(self):
        dstpath = os.path.join(os.getcwd(), f"Files\{self.item_text}")
        
        return dstpath

    def opendialog(self):

            self.filepath =  filedialog.askopenfilename(initialdir="C:/Users/91787", title="select file",
                                            filetypes=(("excel files", "*.xlsx"), ("all files", "*.*")))

            if self.filepath != '':
                self.workbook()
            else:
                print("warning : closed without selection any file")



            
       

# ------------------------ Excel sheet appends --------------------------

    
    def append_datetoexcel(self):
        ws = self.wb[self.month_name]

        
        ws.cell(column = self.length+1 , row =1,  value = self.recent)

        self.wb.save(self.output_save_file())

    def appendpresent(self, n):
        ws = self.wb[self.month_name]
         
        bname = (self.presentbuttons[n])
        ws.cell(row = n+2 , column = self.length+1,value = "P")

        print(ws.cell(row = n+2 , column = 1).value )
        print(ws.cell(row = n+2 , column = self.length+1).value )

        self.wb.save(self.output_save_file())

    def appendabsent(self, n):
        ws = self.wb[self.month_name]

        bname = (self.absentbuttons[n])
        ws.cell(row = n+2 , column = self.length+1, value = 'A') 
        print(ws.cell(row = n+2 , column = 1).value )
        print(ws.cell(row = n+2 , column = self.length+1).value )

        self.wb.save(self.output_save_file())



# ------------------------- sections ----------------------------------------

    def add_tabs(self):

        tab_names = []
        
        for i in self.notebook.tabs():
            tab_names.append(self.notebook.tab(i, "text"))

        if self.item_text in tab_names:
            index = tab_names.index(self.item_text)
            self.notebook.select(index)

            self.tab_frame = self.tab_frame_dict[self.item_text]
            self.wb = load_workbook(self.output_save_file()) 
            self.wordslist()

              
        else:            
            self.tab_frame = Frame(self.notebook, bg = "white")
            self.notebook.add(self.tab_frame, text = self.item_text)
            self.notebook.select(self.tab_frame)

            self.tab_frame_dict[self.item_text] = self.tab_frame




    def loadprevious(self):

        current_path = os.getcwd()
        check = f"{current_path}\\Files\\{self.item_text}\\final.xlsx"
        b = os.path.exists(check)
    
        if b == True:
            self.wb = load_workbook(self.output_save_file()) 
            self.wordslist()
        else:
            self.mb_loadprev()


    def plot_graph(self, n):
        
        ws = self.wb[self.month_name]
       
        total_present = 0
        total_absent = 0

        for i in range(2, self.length + 2):
            value = ws.cell(row = n+2 , column = i).value

            if value == "P":
                total_present += 1
                
            else:
                total_absent += 1
        
        name = ws.cell(row = n+2 , column = 1).value

        # print(f'{name} present : {total_present} times')
        # print(f'{name} absent : {total_absent} times')

        plt.title("Attendance Comparison\nPlot")
        x = ["absent"]
        y = [total_absent]

        x2 = ["present"]
        y2 = [total_present]
        
        
        plt.bar(x,y, label = 'Bar1', color = 'r')
        plt.bar(x2,y2, label = 'Bar2', color = 'c')
        plt.xlabel(f"student: @{name} ")
        plt.ylabel("attendance")
        plt.subplots_adjust(left = 0.226, right = 0.605)
        
        
        self.wb.save(self.output_save_file())
        plt.show()

       
# ------------------------- Excel sheet Extractions ----------------------------------------

    def workbook(self):

        # print(self.get_filepath())
        
        # print(self.destination_path())
        # print(self.output_save_file())
        # print(self.input_save_file())


        if os.path.exists(self.input_save_file()):
            print("Destination file exists!")

            self.wb = load_workbook(self.output_save_file())
            
            
        else:
            shutil.copy(self.filepath , self.destination_path())
            
            self.wb = load_workbook(self.input_save_file())
        
        # self.create_sheet()
        # self.append_datetoexcel()
        self.wordslist()
        
        
    def wordslist(self):

        words = []

        ws = self.wb.active 

        self.row = ws.max_row
        self.column = ws.max_column

        # print(f'maximum row: {self.row}')
        # print(f'maximum column: {self.column}')

        column_a = ws['A']
        

        for i in range(1,self.row):

            words.append(column_a[i].value)
        

        # print(f'list of students: {words}')

        
        self.worksheet_controller(words)



    def worksheet_controller(self, words):

            today = self.today     
            self.sheet_no = self.get_month() 
            self.month_name = today.strftime('%b')
            sheet_no = self.sheet_no 
            month_name = self.month_name

            ws_names = self.wb.sheetnames

            if month_name in ws_names:
                pass
             
            else:

                ws = self.wb.create_sheet(month_name, sheet_no-1)
                ws.cell(column = 1, row = 1, value = "Names' student")

                for i in range(1, self.row):
                    ws.cell(column = 1, row = i+1, value = words[i-1])
                
                self.wb.save(self.output_save_file())

            self.append_datetoexcel()
            self.mainprogram(words)


#   ________________________Interface___________________________


    def Menu(self):

        file_menu = Menu(self.menu, tearoff = 0)
        self.menu.add_cascade(label = "File", menu = file_menu)
        file_menu.add_command(label = 'New File             Ctrl+N ' )
        file_menu.add_separator()
        file_menu.add_command(label = 'Exit' , command = self.master.quit)

       

    def Notebook(self):

        self.notebook.place(x =200 , y = 0 , height = 500 , width = 550)
        


    def sidepanel(self):

        side_panel = Frame(root ,bg = "gray65")
        side_panel.place(x = 750 , y = 0, width = 190 , height = 500)

        
        datelabel = Label(side_panel, text = f'Today is {self.recent}',bg = "black", fg = "white")
        datelabel.grid(row = 8 , column =1, padx =10, pady =1, ipadx = 2, ipady =2 )


        excelbutton = ttk.Button(side_panel, text = 'Excel File', width = 25, command = self.opendialog)
        excelbutton.grid(row = 0 , column = 1,padx = 10, pady = (10,0))

        insert_tree = ttk.Button(side_panel, text = 'CHATBOT', width = 25, command = self.my_bot)
        insert_tree.grid(row = 1 , column = 1,padx = 10, pady = 2)

        load_previous  = ttk.Button(side_panel, text = 'Load Previous', width = 25, command = self.loadprevious)
        load_previous.grid(row = 2 , column = 1,padx = 10, pady = 0)


        exit_button = ttk.Button(side_panel, text = 'Exit', width = 25, command = self.master.quit)
        exit_button.grid(row = 3, column = 1,padx = 10, pady = 50)


    def mainprogram(self, words):

        count = 0
        self.n = 0
        for x in words: 

            count += 1
            jk =  x
            label = Label(self.tab_frame,text=jk, font = ('arial', 10), width = 25) 
            label.grid(row = count, column = 0, pady = 4 )

            presentbutton = ttk.Button(self.tab_frame,text = "present", command = partial(self.appendpresent , self.n))
            presentbutton.grid(row = count , column = 3 , padx = (50,0) )
            
            absentbutton = ttk.Button(self.tab_frame,text = "absent", command = partial(self.appendabsent , self.n))
            absentbutton.grid(row = count , column = 4, padx = 4)

            graphbutton = ttk.Button(self.tab_frame, text = "Graph", command = partial(self.plot_graph, self.n))
            graphbutton.grid(row = count, column = 5, padx = 4)
                
            self.labels.append(label) #appends the label to the list for further use
            self.presentbuttons.append(presentbutton)
            self.absentbuttons.append(absentbutton)
            self.n += 1




    def doitall(self):
        self.dumpdatelist()
        self.Menu()
        self.Treeview()
        self.Notebook()
        self.sidepanel()






root = Tk()
root.geometry('940x500+320+110')
root.resizable(False, False)

app = application(root)
app.doitall()

root.mainloop()